import openpyxl
import pandas as pd
import numpy as np
import re


def get_excel_sheet_name_list(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet_list = wb.sheetnames
    return sheet_list


def excel_convert_to_text(file_name, sheet_name):

    wb = openpyxl.load_workbook(file_name)
    sheet = wb[sheet_name]

    df = pd.read_excel(file_name, sheet_name=sheet_name, header=None)
    df = df.fillna(' ')  # 空格填补nan值
    cells = df.to_numpy().astype(str)  # 转换成字符串np格式

    r, c = cells.shape
    bd = np.zeros((r, c), dtype='int16')  # 是否加粗
    an = np.zeros((r, c), dtype=np.str_)  # 对齐方式
    stp = np.zeros((r, c), dtype='int16')  # 单元格上边框
    sbm = np.zeros((r, c), dtype='int16')  # 单元格下边框
    slt = np.zeros((r, c), dtype='int16')  # 单元格左边框
    srt = np.zeros((r, c), dtype='int16')  # 单元格右边框
    bcell = np.zeros((r, c), dtype=np.dtype('U100'))  # 带加粗信息单元格内容

    for i in range(1, r+1):  # 读取单元格格式
        for j in range(1, c+1):
            ib = sheet.cell(row=i, column=j).font.bold
            ia = sheet.cell(row=i, column=j).alignment.horizontal
            istp = sheet.cell(row=i, column=j).border.top.style
            isbm = sheet.cell(row=i, column=j).border.bottom.style
            islt = sheet.cell(row=i, column=j).border.left.style
            isrt = sheet.cell(row=i, column=j).border.right.style
            bd[i - 1][j - 1] = ib
            an[i - 1][j - 1] = ia
            stp[i - 1][j - 1] = (istp != None)
            sbm[i - 1][j - 1] = (isbm != None)
            slt[i - 1][j - 1] = (islt != None)
            srt[i - 1][j - 1] = (isrt != None)
            if ib == 1:  # 写入加粗指令
                bcell[i - 1][j - 1] = '\\bf {{{0}}}'.format(cells[i-1][j-1])
            else:
                bcell[i - 1][j - 1] = cells[i-1][j-1]

    mc = sheet.merged_cells.ranges  # 获取合并单元格列表
    mcnp = np.zeros((len(mc), 4), dtype='str_')
    mm = ','.join(map(str, mc))
    mm = mm.split(',')

    count = 0
    for ele in mm:  # 查找每个合并单元格范围
        count = count+1
        match = re.search(r'([A-Z]\d+):([A-Z]\d+)', ele)  # 限定单元格列索引A-Z
        start_col = match.group(1)[0]
        end_col = match.group(2)[0]
        start_row = match.group(1)[1:]
        end_row = match.group(2)[1:]
        mcnp[count - 1][0] = start_col  # 合并单元格的起始行列提取
        mcnp[count - 1][1] = start_row
        mcnp[count - 1][2] = end_col
        mcnp[count - 1][3] = end_row

    for merge in mcnp:  # 合并单元格框线合并
        r0 = int(merge[1]) - 1
        r1 = int(merge[3]) - 1
        c0 = ord(merge[0]) - ord('A')
        c1 = ord(merge[2]) - ord('A')
        for i in range(c0, c1+1):
            for j in range(r0, r1+1):
                if stp[j][i] == 1:
                    stp[r0:r1+1, i] = 0
                    stp[r0, i] = 1

                if sbm[j][i] == 1:
                    sbm[r0:r1+1, i] = 0
                    sbm[r1, i] = 1

                if slt[j][i] == 1:
                    slt[r0:r1+1, c0:c1+1] = 0
                    slt[r0:r1+1, c0] = 1

                if srt[j][i] == 1:
                    srt[r0:r1+1, c0:c1+1] = 0
                    srt[r0:r1+1, c1] = 1
                    break

    hsidenp = np.zeros((r+1, c), dtype='int16')  # 表格水平框线
    hsidenp[0] = stp[0]
    for i in range(0, r-1):
        hsidenp[i+1] = sbm[i] | stp[i+1]
    hsidenp[r] = sbm[r-1]

    vsidenp = np.zeros((r, c+1), dtype='int16')  # 表格竖直框线
    vsidenp[:, 0] = slt[:, 0]
    for i in range(0, c-1):
        vsidenp[:, i+1] = srt[:, i] | slt[:, i+1]
    vsidenp[:, c] = srt[:, c-1]

    for merge in mcnp:  # multicolumn和multirow处理
        r0 = int(merge[1]) - 1
        r1 = int(merge[3]) - 1
        c0 = ord(merge[0]) - ord('A')
        c1 = ord(merge[2]) - ord('A')
        if r1 - r0 >= 1:
            bcell[r0][c0] = '\\multirow{{{0}}}{{*}}{{{1}}}'.format(r1 - r0 + 1, bcell[r0][c0])
        if c1 - c0 >= 1:
            if vsidenp[r0][c0] == 1:
                vl = '|'
            else:
                vl = ''
            if vsidenp[r0][c1+1] == 1:
                vr = '|'
            else:
                vr = ''
            bcell[r0][c0] = '\\multicolumn{{{0}}}{{{1}c{2}}}{{{3}}}'.format(c1-c0+1, vl, vr, bcell[r0][c0])

    #  开始和结尾固定格式
    tabular = '\t\t\\begin{tabular}'+'{'
    text = ['\\begin{table}[htbp]', '\t\\caption{}', '\t\\begin{center}']
    endtext = ['\t\t\\end{tabular}', '\t\t\\label{}', '\t\\end{center}', '\\end{table}']

    for i in range(0, c+1):  # 单元格对齐格式设置和表格竖框线设置
        if 1 in vsidenp[:, i]:
            tabular = tabular + '|'
        else:
            tabular = tabular + ''
        if i < c:
            tabular = tabular + 'c'
    tabular = tabular + '}'
    text.append(tabular)

    hsidecmd = []
    for i in range(0, r+1):  # 横框线指令
        vcount = 0
        for j in range(0, c):
            if hsidenp[i][j] == 1:
                vcount = vcount + 1
                if vcount == 1:  # 横框线首次出现位置
                    hidx = j
        if vcount == c:
            hsy = '\\hline'
        elif vcount == 0:
            hsy = ''
        else:
            hsy = '\\cline{{{}-{}}}'.format(hidx + 1, hidx + vcount)
        hsidecmd.append(hsy)
    text.append('\t\t' + hsidecmd[0])  # 表格整体上框线

    for i in range(0, r):  # 处理每行单元格内容
        res = []
        htext = '\t\t'
        for j in range(0, c):
            res.append(bcell[i][j])
        for merge in mcnp:  # 删去multicolumn剩下空内容
            r0 = int(merge[1]) - 1
            r1 = int(merge[3]) - 1
            c0 = ord(merge[0]) - ord('A')
            c1 = ord(merge[2]) - ord('A')
            if i >= r0 and i <= r1:
                if c1 - c0 >= 1:
                    del res[c0+1:c1+1]  # 合并单元格除第一个单元格的剩下内容，默认multicolumn只有一行
        htext = htext + '& '.join(res) + '\\\\' + hsidecmd[i+1]  # 每行内容的指令
        text.append(htext)

    text = text + endtext  # 总体指令合并
    file = open('tex.txt', 'w')  # 输出指令到当前文件夹
    for line in text:
        file.write(line + '\n')
    file.close()

